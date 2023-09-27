import requests
import pandas as pd
from requests.exceptions import RequestException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import os

API_TOKEN = "YOUR_API_KEY_HERE"
PODCAST_URL = "YOUR_EPISODES_URL_HERE"
EXCEL_FILE_PATH = "YOUR_PATH_HERE"

headers = {"Authorization": f"Token token={API_TOKEN}"}

session = requests.Session()
session.headers.update(headers)

try:
    response = session.get(PODCAST_URL)
    response.raise_for_status()
except RequestException as e:
    print(f"Er is een fout opgetreden bij het ophalen van de afleveringen: {e}")
else:
    try:
        episodes = response.json()
    except ValueError:
        print("De ontvangen respons is geen geldige JSON.")
    else:
        stats_list = []

        for episode in episodes:
            episode_id = episode.get('id')
            episode_title = episode.get('title')
            episode_number = episode.get('episode_number')
            audio_url = episode.get('audio_url')
            artwork_url = episode.get('artwork_url')
            magic_mastering = episode.get('magic_mastering')

            published_at = episode.get('published_at')
            if published_at is not None:
                published_at = pd.to_datetime(published_at).date()

            duration = episode.get('duration')
            if duration is not None:
                duration = f"{duration // 3600} uur {duration % 3600 // 60} min"

            stats_url = f"YOUR_API_URL_HERE"

            try:
                stats_response = session.get(stats_url)
                stats_response.raise_for_status()
            except RequestException as e:
                print(f"Kon geen statistieken ophalen voor aflevering {episode_id}. Fout: {e}")
            else:
                try:
                    stats_data = stats_response.json()
                    total_plays = stats_data.get('total_plays')
                    if total_plays is None:
                        total_plays = ""

                    stats_list.append({
                        'Afleveringsnummer': episode_number,
                        'Afleveringsidentificatie': episode_id,
                        'Afleveringstitel': episode_title,
                        'Audio URL': audio_url,
                        'Artwork URL': artwork_url,
                        'Gepubliceerd op': published_at,
                        'Totale afspeelduur': duration,
                        'Magic Mastering': magic_mastering,
                        'Totaal afspelen': total_plays,
                    })
                except ValueError:
                    print(f"Geen geldige JSON in de reactie voor aflevering {episode_id}")

        if stats_list:
            df = pd.DataFrame(stats_list)
            df = df.sort_values(by='Afleveringsnummer', ascending=True)

            try:
                if not os.path.exists(EXCEL_FILE_PATH):
                    workbook = Workbook()
                    worksheet = workbook.active
                    worksheet.title = "Statistieken"
                    include_header = True
                else:
                    workbook = load_workbook(EXCEL_FILE_PATH)
                    if 'Statistieken' not in workbook.sheetnames:
                        worksheet = workbook.create_sheet(title='Statistieken')
                        include_header = True
                    else:
                        worksheet = workbook['Statistieken']
                        include_header = worksheet.max_row == 1 

                if worksheet.max_row > 1:
                    worksheet.delete_rows(2, worksheet.max_row)

                for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
                    for j, value in enumerate(row, start=1):
                        worksheet.cell(row=i, column=j, value=value)

                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                for column_cells in worksheet.columns:
                    max_length = 0
                    column = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except TypeError:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column].width = adjusted_width

                workbook.save(EXCEL_FILE_PATH)
                print(f"Statistieken zijn opgeslagen in het Excel-bestand: {EXCEL_FILE_PATH}")
            except Exception as e:
                print(f"Fout bij het bijwerken van het Excel-bestand: {e}")
        else:
            print("Er waren geen geldige afleveringsstatistieken om op te slaan.")
